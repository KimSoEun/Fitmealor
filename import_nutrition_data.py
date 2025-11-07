"""
Import official nutrition data from Excel files to SQLite database
식품의약품안전처 공식 영양 데이터를 엑셀 파일에서 SQLite 데이터베이스로 가져오기
"""
import sqlite3
import json
from openpyxl import load_workbook
import sys

DB_PATH = "fitmealor.db"

# Excel files to import
EXCEL_FILES = [
    "/Users/goorm/Fitmealor/backend/data/20250327_가공식품DB_147999건.xlsx",
    "/Users/goorm/Fitmealor/backend/data/20250408_음식DB.xlsx"
]

def parse_allergens(food_name, food_category):
    """
    Extract potential allergens from food name and category
    식품명과 카테고리에서 잠재적 알레르기 유발 물질 추출
    """
    allergens = []

    # Common Korean allergen keywords
    allergen_map = {
        'eggs': ['계란', '달걀', '에그', '난', '마요네즈'],
        'dairy': ['우유', '치즈', '요거트', '요구르트', '크림', '버터', '유제품'],
        'peanuts': ['땅콩', '피넛'],
        'tree_nuts': ['아몬드', '호두', '캐슈', '피스타치오', '밤', '잣'],
        'fish': ['생선', '어류', '참치', '연어', '고등어', '명태', '대구'],
        'shellfish': ['새우', '게', '랍스터', '조개', '굴', '홍합', '해산물'],
        'soy': ['콩', '두부', '된장', '간장', '낫토'],
        'wheat': ['밀', '밀가루', '빵', '면', '파스타', '우동'],
    }

    food_text = (food_name + ' ' + food_category).lower()

    for allergen_key, keywords in allergen_map.items():
        for keyword in keywords:
            if keyword in food_text:
                allergens.append(allergen_key)
                break

    return allergens


def import_from_excel(file_path):
    """
    Import meals from Excel file to database
    엑셀 파일에서 식사 데이터를 데이터베이스로 가져오기
    """
    print(f"\n{'='*80}")
    print(f"📂 Processing: {file_path.split('/')[-1]}")
    print(f"{'='*80}")

    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    # Get headers from first row
    headers = [cell.value for cell in ws[1]]

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    imported_count = 0
    skipped_count = 0
    error_count = 0

    # Process each row
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            # Create dictionary from row data
            row_data = dict(zip(headers, row))

            # Extract key fields
            meal_id = row_data.get('식품코드')
            name = row_data.get('식품명')
            category = row_data.get('식품대분류명', '')
            brand = row_data.get('제조사명') or row_data.get('업체명') or ''

            # Skip if no meal_id or name
            if not meal_id or not name:
                skipped_count += 1
                continue

            # Extract nutrition data
            calories = row_data.get('에너지(kcal)')
            protein_g = row_data.get('단백질(g)')
            carbs_g = row_data.get('탄수화물(g)')
            fat_g = row_data.get('지방(g)')
            sodium_mg = row_data.get('나트륨(mg)')

            # Parse serving size
            serving_size = row_data.get('영양성분함량기준량', '100g')

            # Parse allergens from food name and category
            allergens = parse_allergens(str(name), str(category))

            # Extract origin/source
            origin = row_data.get('원산지국명', '')
            if origin == '해당없음' or not origin:
                origin = '국내산'

            # Calculate a simple score based on nutrition
            score = 70  # Base score
            try:
                if calories and protein_g:
                    # Higher protein-to-calorie ratio = better score
                    protein_ratio = float(protein_g) / (float(calories) + 1) * 100
                    if protein_ratio > 20:
                        score += 20
                    elif protein_ratio > 10:
                        score += 10

                    # Lower sodium = better score
                    if sodium_mg:
                        if float(sodium_mg) < 200:
                            score += 10
                        elif float(sodium_mg) > 800:
                            score -= 10
            except:
                pass

            # Ensure score is within 0-100
            score = max(0, min(100, score))

            # Create simple explanation
            explanation_ko = f"{name}. {category}."
            if calories:
                explanation_ko += f" 100g당 {calories}kcal."
            if protein_g:
                explanation_ko += f" 단백질 {protein_g}g."

            explanation_en = f"{name}. Category: {category}."
            if calories:
                explanation_en += f" {calories} kcal per 100g."
            if protein_g:
                explanation_en += f" Protein: {protein_g}g."

            # Insert into database
            cursor.execute("""
                INSERT OR IGNORE INTO meals VALUES (
                    ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?
                )
            """, (
                str(meal_id),
                str(name),
                str(name),  # name_en (same as Korean name for now)
                str(brand) if brand else '공식 영양 DB',
                str(category),
                json.dumps([], ensure_ascii=False),  # ingredients (not available)
                json.dumps(allergens, ensure_ascii=False),
                int(float(calories)) if calories else None,
                float(protein_g) if protein_g else None,
                float(carbs_g) if carbs_g else None,
                float(fat_g) if fat_g else None,
                int(float(sodium_mg)) if sodium_mg else None,
                str(serving_size),
                str(origin),
                explanation_en,
                explanation_ko,
                int(score)
            ))

            imported_count += 1

            # Progress update every 10000 rows
            if imported_count % 10000 == 0:
                print(f"   ✅ Imported {imported_count} items...")
                conn.commit()  # Commit in batches

        except Exception as e:
            error_count += 1
            if error_count <= 5:  # Only print first 5 errors
                print(f"   ⚠️ Row {row_idx} error: {e}")
            continue

    # Final commit
    conn.commit()
    conn.close()
    wb.close()

    print(f"\n📊 Import Summary:")
    print(f"   ✅ Imported: {imported_count}")
    print(f"   ⏭️  Skipped: {skipped_count}")
    print(f"   ❌ Errors: {error_count}")

    return imported_count


def verify_import():
    """Verify the imported data"""
    print(f"\n{'='*80}")
    print("🔍 Verifying imported data...")
    print(f"{'='*80}")

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    # Total count
    cursor.execute("SELECT COUNT(*) FROM meals")
    total_count = cursor.fetchone()[0]
    print(f"\n📊 Total meals in database: {total_count:,}")

    # Sample by category
    cursor.execute("""
        SELECT category, COUNT(*) as count
        FROM meals
        GROUP BY category
        ORDER BY count DESC
        LIMIT 10
    """)
    print(f"\n📋 Top 10 categories:")
    for category, count in cursor.fetchall():
        print(f"   - {category}: {count:,}")

    # Sample high-protein meals
    cursor.execute("""
        SELECT name, calories, protein_g, score
        FROM meals
        WHERE protein_g > 20
        ORDER BY score DESC
        LIMIT 5
    """)
    print(f"\n🥇 Top 5 high-protein meals:")
    for name, calories, protein, score in cursor.fetchall():
        print(f"   - {name}: {calories}kcal, {protein}g protein, score: {score}")

    conn.close()


def main():
    """Main import process"""
    print("🚀 Starting nutrition data import...")
    print(f"📁 Files to process: {len(EXCEL_FILES)}")

    # Check if database already has data
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) FROM meals")
    existing_count = cursor.fetchone()[0]
    conn.close()

    if existing_count > 0:
        response = input(f"\n⚠️  Database already contains {existing_count} meals. Clear and re-import? (y/n): ")
        if response.lower() == 'y':
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute("DELETE FROM meals")
            conn.commit()
            conn.close()
            print("🗑️  Cleared existing meal data.")
        else:
            print("❌ Import cancelled.")
            return

    total_imported = 0

    # Import from each Excel file
    for file_path in EXCEL_FILES:
        try:
            imported = import_from_excel(file_path)
            total_imported += imported
        except Exception as e:
            print(f"❌ Failed to import {file_path}: {e}")
            import traceback
            traceback.print_exc()

    print(f"\n{'='*80}")
    print(f"✅ Total imported across all files: {total_imported:,}")
    print(f"{'='*80}")

    # Verify the import
    verify_import()

    print("\n✅ Import completed successfully!")


if __name__ == "__main__":
    main()
